from rest_framework import viewsets, filters, status
from rest_framework.decorators import action
from rest_framework.response import Response
from django_filters.rest_framework import DjangoFilterBackend
from core.permissions import CanManageProducts
from .models import Category, Product, InventoryAdjustment
from .serializers import CategorySerializer, ProductSerializer, InventoryAdjustmentSerializer

from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db.models import Q
from django.core.paginator import Paginator

from .models import Product, Category
from .forms import ProductForm, CategoryForm

class CategoryViewSet(viewsets.ModelViewSet):
    queryset = Category.objects.all()
    serializer_class = CategorySerializer
    permission_classes = [CanManageProducts]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['parent', 'is_active']
    search_fields = ['name']
    ordering_fields = ['name', 'display_order']

class ProductViewSet(viewsets.ModelViewSet):
    queryset = Product.objects.select_related('category').all()
    serializer_class = ProductSerializer
    permission_classes = [CanManageProducts]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['category', 'is_active', 'track_stock']
    search_fields = ['sku', 'barcode', 'name']
    ordering_fields = ['name', 'sell_price', 'stock']
    
    @action(detail=False, methods=['get'])
    def low_stock(self, request):
        """Get products with low stock"""
        low_stock_products = [p for p in self.queryset if p.is_low_stock]
        serializer = self.get_serializer(low_stock_products, many=True)
        return Response(serializer.data)
    
    @action(detail=True, methods=['post'])
    def adjust_stock(self, request, pk=None):
        """Adjust product stock"""
        product = self.get_object()
        quantity = request.data.get('quantity', 0)
        reason = request.data.get('reason', '')
        
        try:
            product.adjust_stock(
                quantity=int(quantity),
                reason=reason,
                performed_by=request.user
            )
            return Response({'status': 'stock adjusted'})
        except Exception as e:
            return Response(
                {'error': str(e)},
                status=status.HTTP_400_BAD_REQUEST
            )

class InventoryAdjustmentViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = InventoryAdjustment.objects.select_related('product', 'performed_by').all()
    serializer_class = InventoryAdjustmentSerializer
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter]
    filterset_fields = ['product', 'reason']
    ordering_fields = ['timestamp']


@login_required
def product_list(request):
    """List all products"""
    # Search and filter
    search_query = request.GET.get('search', '')
    category_filter = request.GET.get('category', '')
    status_filter = request.GET.get('status', '')
    
    products = Product.objects.select_related('category').all()
    
    if search_query:
        products = products.filter(
            Q(sku__icontains=search_query) |
            Q(name__icontains=search_query) |
            Q(barcode__icontains=search_query)
        )
    
    if category_filter:
        products = products.filter(category_id=category_filter)
    
    if status_filter == 'active':
        products = products.filter(is_active=True)
    elif status_filter == 'inactive':
        products = products.filter(is_active=False)
    elif status_filter == 'low_stock':
        products = [p for p in products if p.is_low_stock]
    
    # Pagination
    paginator = Paginator(products, 50)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    categories = Category.objects.filter(is_active=True)
    
    context = {
        'page_obj': page_obj,
        'categories': categories,
        'search_query': search_query,
        'category_filter': category_filter,
        'status_filter': status_filter,
    }
    
    return render(request, 'catalog/product_list.html', context)


@login_required
def product_detail(request, pk):
    """View product details"""
    product = get_object_or_404(Product.objects.select_related('category'), pk=pk)
    
    # Get recent inventory adjustments
    recent_adjustments = product.adjustments.select_related('performed_by').order_by('-timestamp')[:10]
    
    context = {
        'product': product,
        'recent_adjustments': recent_adjustments,
    }
    
    return render(request, 'catalog/product_detail.html', context)


@login_required
def product_create(request):
    """Create new product"""
    if not request.user.can_manage_products:
        messages.error(request, "You don't have permission to create products.")
        return redirect('catalog:product_list')
    
    if request.method == 'POST':
        form = ProductForm(request.POST, request.FILES)
        if form.is_valid():
            product = form.save(commit=False)
            product.created_by = request.user
            product.save()
            
            from core.models import AuditLog
            AuditLog.log(
                user=request.user,
                action=AuditLog.Action.CREATE_PRODUCT,
                description=f"Created product {product.sku} - {product.name}"
            )
            
            messages.success(request, f'Product "{product.name}" created successfully.')
            return redirect('catalog:product_detail', pk=product.pk)
    else:
        form = ProductForm()
    
    context = {
        'form': form,
        'action': 'Create',
    }
    
    return render(request, 'catalog/product_form.html', context)


@login_required
def product_update(request, pk):
    """Update product"""
    if not request.user.can_manage_products:
        messages.error(request, "You don't have permission to update products.")
        return redirect('catalog:product_list')
    
    product = get_object_or_404(Product, pk=pk)
    
    if request.method == 'POST':
        form = ProductForm(request.POST, request.FILES, instance=product)
        if form.is_valid():
            form.save()
            
            from core.models import AuditLog
            AuditLog.log(
                user=request.user,
                action=AuditLog.Action.UPDATE_PRODUCT,
                description=f"Updated product {product.sku} - {product.name}"
            )
            
            messages.success(request, f'Product "{product.name}" updated successfully.')
            return redirect('catalog:product_detail', pk=product.pk)
    else:
        form = ProductForm(instance=product)
    
    context = {
        'form': form,
        'product': product,
        'action': 'Update',
    }
    
    return render(request, 'catalog/product_form.html', context)


@login_required
def product_delete(request, pk):
    """Delete product"""
    if not request.user.can_manage_products:
        messages.error(request, "You don't have permission to delete products.")
        return redirect('catalog:product_list')
    
    product = get_object_or_404(Product, pk=pk)
    
    if request.method == 'POST':
        product_name = product.name
        product.delete()
        
        from core.models import AuditLog
        AuditLog.log(
            user=request.user,
            action=AuditLog.Action.DELETE_PRODUCT,
            description=f"Deleted product {product.sku} - {product_name}"
        )
        
        messages.success(request, f'Product "{product_name}" deleted successfully.')
        return redirect('catalog:product_list')
    
    context = {
        'product': product,
    }
    
    return render(request, 'catalog/product_confirm_delete.html', context)


@login_required
def product_export(request):
    """Export products to Excel"""
    if not request.user.can_manage_products:
        messages.error(request, "You don't have permission to export products.")
        return redirect('catalog:product_list')
    
    import openpyxl
    from django.http import HttpResponse
    from datetime import datetime
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename=products_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = 'Products'
    
    # Header row
    columns = [
        'SKU', 'Barcode', 'Name', 'Category', 'Cost Price', 'Sell Price', 
        'Tax Rate', 'Stock', 'Low Stock Threshold', 'Unit', 'Is Active', 'Track Stock'
    ]
    worksheet.append(columns)
    
    # Data rows
    products = Product.objects.select_related('category').all()
    for product in products:
        worksheet.append([
            product.sku,
            product.barcode,
            product.name,
            product.category.name if product.category else '',
            product.cost_price,
            product.sell_price,
            product.tax_rate,
            product.stock,
            product.low_stock_threshold,
            product.unit,
            product.is_active,
            product.track_stock
        ])
    
    workbook.save(response)
    return response


@login_required
def product_import(request):
    """Import products from Excel"""
    if not request.user.can_manage_products:
        messages.error(request, "You don't have permission to import products.")
        return redirect('catalog:product_list')
    
    from .forms import ProductImportForm
    import openpyxl
    
    if request.method == 'POST':
        form = ProductImportForm(request.POST, request.FILES)
        if form.is_valid():
            excel_file = request.FILES['excel_file']
            try:
                workbook = openpyxl.load_workbook(excel_file)
                worksheet = workbook.active
                
                # Skip header row
                rows = worksheet.iter_rows(min_row=2, values_only=True)
                
                created_count = 0
                updated_count = 0
                errors = []
                
                for i, row in enumerate(rows, start=2):
                    try:
                        # Basic validation - check required fields
                        sku = row[0]
                        name = row[2]
                        
                        if not sku or not name:
                            errors.append(f"Row {i}: SKU and Name are required.")
                            continue
                            
                        # Get or create category
                        category_name = row[3]
                        category = None
                        if category_name:
                            category, _ = Category.objects.get_or_create(name=category_name)
                        
                        # Update or create product
                        product, created = Product.objects.update_or_create(
                            sku=sku,
                            defaults={
                                'barcode': row[1] or '',
                                'name': name,
                                'category': category,
                                'cost_price': row[4] or 0,
                                'sell_price': row[5] or 0,
                                'tax_rate': row[6] or 0.15,
                                'stock': row[7] or 0,
                                'low_stock_threshold': row[8] or 10,
                                'unit': row[9] or 'piece',
                                'is_active': row[10] if row[10] is not None else True,
                                'track_stock': row[11] if row[11] is not None else True,
                            }
                        )
                        
                        if created:
                            created_count += 1
                        else:
                            updated_count += 1
                            
                    except Exception as e:
                        errors.append(f"Row {i}: {str(e)}")
                
                if errors:
                    for error in errors[:5]:  # Show first 5 errors
                        messages.warning(request, error)
                    if len(errors) > 5:
                        messages.warning(request, f"And {len(errors) - 5} more errors.")
                
                messages.success(request, f"Import complete: {created_count} created, {updated_count} updated.")
                
                from core.models import AuditLog
                AuditLog.log(
                    user=request.user,
                    action=AuditLog.Action.CREATE_PRODUCT,
                    description=f"Imported products: {created_count} created, {updated_count} updated"
                )
                
            except Exception as e:
                messages.error(request, f"Error processing file: {str(e)}")
            
            return redirect('catalog:product_list')
    else:
        form = ProductImportForm()
        
    # If GET request (or invalid form), we don't render a separate page for import
    # The modal is in the product list page. This view is primarily for POST.
    # But if accessed directly, redirect to list.
    return redirect('catalog:product_list')